import  openpyxl

# wb = openpyxl.Workbook()
# ws = wb.active
# # 更改默认名称Sheet`
# ws.title = 'WorkSheetTitle'
# ws2 = wb.create_sheet('WorkSheetTitle_2')
# ws3 = wb.create_sheet('WorkSheetTitle_3',1)
# 给单元格赋值
# ws["A1"] = "项目模块"
# ws["B1"] = "模块"
# ws["C1"] = "时间"
# 指定行列给单元格赋值
# ws.cell(row=4, column=2, value=10)
# 指定行列给单元格赋值
# v = 0
# for i in range(1, 10):
#     for n in range(1, 10):
#         ws.cell(row=i, column=n, value=v)
#         v += 1

# 单元格内换行
# ws['A1'] = "A\nB\nC"
# ws['A1'].alignment = openpyxl.styles.Alignment(wrapText=True)

# wb.save('user.xlsx')

# column名
# column_title = ["FirstName", "LastName"]
#
# if __name__ == '__main__':
#     """
#     CELL放入值
#     """
#     wb = openpyxl.Workbook()
#     ws = wb.active
#
#     # 更改默认名称Sheet`
#     ws.title = "worksheettitle"
#
#     # column名和値顺序放入单元格中
#     rows = [
#         column_title,
#         ["Tarou", "Tanaka"],
#         ["Tarou", "Suzuki"],
#         ["Tarou", "Uchiayama"],
#     ]
#     for row in rows:
#         ws.append(row)

    # 保存
    # wb.save('example.xlsx')
wb = openpyxl.load_workbook('example.xlsx')
sheet_name = wb.get_sheet_names()
print('该表格总有%d个sheet，分别为：%s'%(len(sheet_name),sheet_name))
